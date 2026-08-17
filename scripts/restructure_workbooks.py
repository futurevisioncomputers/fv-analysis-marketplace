"""Restructure the institute's two workbooks into linked, referential sheets.

Shape (the operator's, with one master added):

  FV_Enquiry.xlsx      enquiries (PK ENQ_ID) -> follow_up_calls (FK ENQ_ID)
  FV_Students.xlsx     students  (PK student_id) -> fee_receipts, certificates

Interlinking is done two different ways ON PURPOSE:

  * WITHIN a workbook, child sheets carry live XLOOKUP formulas back to their
    parent, and every closed-set column gets a dropdown pointing at `lookups`.
    A name or a branch is therefore stored in exactly one place.
  * ACROSS the two workbooks, only the `student_id` VALUE is copied onto the
    enquiry row. A cross-file formula resolves to a full path and breaks the
    moment the other file is moved, renamed, or closed — the bridge has to
    survive that, so it is a typed value, not a link.

Nothing is invented: a status that cannot be derived is written as Unknown,
and the counts are printed so the manual work left is visible.
"""

from __future__ import annotations

import argparse
import os
import re
import sys

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# The repo root, from this file's own location — not a machine's absolute path.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from agents import canonical_maps as cm  # noqa: E402

# Where the exported CSVs live and where the workbooks are written. Both are
# arguments; the defaults are repo-relative so this runs on a fresh checkout.
REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(REPO, "data", "sheets")
OUT = os.path.join(REPO, "output")

# Lifecycle precedence when a student carries more than one status row. Most
# advanced state wins: having completed is not undone by an older "learning".
PRECEDENCE = ["Course Completed", "Not To Entertain", "Not Coming",
              "Currently Learning"]


def read(name: str, src: str = "") -> pd.DataFrame:
    path = os.path.join(src or SRC, f"{name}.csv")
    if not os.path.exists(path):
        raise SystemExit(
            f"no {name}.csv in {src or SRC}\n"
            f"Point --src at the directory holding the exported sheets.")
    return pd.read_csv(path)


def as_id(series: pd.Series) -> pd.Series:
    """student_id arrives as float in one sheet and int in another."""
    return pd.to_numeric(series, errors="coerce").astype("Int64")


def build_students(adm, st, fee, cert):
    """One row per student, from every sheet that knows one exists."""
    st = st.copy()
    for df, col in ((adm, "student_id"), (st, "student_id"),
                    (fee, "student_id"), (cert, "student_id")):
        df[col] = as_id(df[col])

    ids = pd.Index(
        sorted(set(adm["student_id"].dropna()) | set(st["student_id"].dropna())
               | set(fee["student_id"].dropna()) | set(cert["student_id"].dropna())),
        name="student_id")
    students = pd.DataFrame(index=ids).reset_index()

    # Attributes come from admissions first (richest), then certificates for
    # the historical students who have no admission row at all.
    a = adm.drop_duplicates("student_id").set_index("student_id")
    c = cert.drop_duplicates("student_id").set_index("student_id")

    def pick(col_a, col_c=None):
        left = students["student_id"].map(a[col_a]) if col_a in a.columns else pd.Series(index=students.index, dtype=object)
        if col_c and col_c in c.columns:
            right = students["student_id"].map(c[col_c])
            return left.where(left.notna(), right)
        return left

    students["name"] = pick("student_name", "name")
    students["mobile"] = pick("mobile_clean")
    students["course"] = pick("course", "course")
    students["branch"] = pick("branch", "branch")
    students["faculty"] = pick("faculty")
    students["admission_date"] = pick("admission_date")
    students["joining_date"] = pick("admission_date", "date_of_joining")
    students["total_fees"] = pick("total_fees_actual")
    students["fee_status"] = pick("fee_status")
    students["amt_pending"] = pick("amt_pending")
    students["presently_doing"] = pick("presently_doing")
    students["lead_source"] = pick("lead_source_bucket")

    # --- reporting taxonomies (the institute's closed sets) ---------------
    students["course_category"] = [
        cm.canonicalize_course_category(v) for v in students["course"]]
    students["presently_doing"] = students["presently_doing"].map(
        cm.canonicalize_occupation)
    students["lead_source"] = students["lead_source"].map(
        cm.canonicalize_lead_source)
    students["branch"] = students["branch"].map(
        lambda v: (cm.branch_from_locality(v) or cm.canonicalize_branch(v)))
    students["branch"] = students["branch"].map(
        lambda v: v.title() if isinstance(v, str) else v)

    # --- status: derived, never guessed ----------------------------------
    st["status"] = st["status"].astype(str).str.strip()
    rank = {s: i for i, s in enumerate(PRECEDENCE)}
    st["_rank"] = st["status"].map(rank).fillna(len(PRECEDENCE))
    best = st.sort_values("_rank").drop_duplicates("student_id").set_index("student_id")

    students["status"] = students["student_id"].map(best["status"])
    students["status_detail"] = students["student_id"].map(best["status_detail"])
    students["days_remaining"] = students["student_id"].map(best["days_remaining"])

    from_status = students["status"].notna().sum()

    # A certificate that has actually been issued is proof of completion.
    has_cert = c["certificate_issue_date"].notna()
    cert_done = students["student_id"].map(has_cert).fillna(False).astype(bool)
    gap = students["status"].isna() & cert_done
    students.loc[gap, "status"] = "Course Completed"
    students.loc[gap, "status_detail"] = "derived: certificate issued"
    from_cert = int(gap.sum())

    students["status"] = students["status"].fillna("Unknown")
    students["status_date"] = ""          # for the operator to fill
    students["course_duration_days"] = ""  # from Student_Time_Table2023
    students["ENQ_ID"] = ""                # the bridge, filled below

    return students, from_status, from_cert


def repair_enquiries(enq, fu, conv, students):
    """Give follow-up calls a real ENQ_ID, and enquiries a student_id."""
    enq = enq.copy()
    fu = fu.copy()
    enq["mobile_clean"] = pd.to_numeric(enq["mobile_clean"], errors="coerce").astype("Int64")
    fu["mobile_clean"] = pd.to_numeric(fu["mobile_clean"], errors="coerce").astype("Int64")

    # ENQ_Number in the source holds junk ('9013', '-', free text). Mobile is
    # the key that actually joins, at 834/834.
    lookup = enq.dropna(subset=["mobile_clean"]).drop_duplicates(
        "mobile_clean").set_index("mobile_clean")["ENQ_ID"]
    fu["ENQ_ID"] = fu["mobile_clean"].map(lookup)
    repaired = int(fu["ENQ_ID"].notna().sum())

    # --- taxonomies on the enquiry side too ------------------------------
    enq["lead_source"] = enq["lead_source_raw"].map(cm.canonicalize_lead_source)
    enq["presently_doing"] = enq["presently_doing"].map(cm.canonicalize_occupation)
    enq["course_category"] = [cm.canonicalize_course_category(v) for v in enq["course"]]
    enq["branch"] = enq["branch"].map(
        lambda v: (cm.branch_from_locality(v) or cm.canonicalize_branch(v)))
    enq["branch"] = enq["branch"].map(lambda v: v.title() if isinstance(v, str) else v)
    enq["counsellor"] = enq["counsellor"].map(cm.canonicalize_faculty)
    enq["counsellor"] = enq["counsellor"].map(
        lambda v: v.title() if isinstance(v, str) else v)

    # The bridge. `converted` matches admissions on mobile (124/169) but NOT
    # enquiries (1/169) — the two workbooks were masked separately — so the
    # link can only be filled where a mobile genuinely matches.
    smap = students.dropna(subset=["mobile"]).copy()
    smap["mobile"] = pd.to_numeric(smap["mobile"], errors="coerce").astype("Int64")
    smap = smap.dropna(subset=["mobile"]).drop_duplicates("mobile").set_index("mobile")["student_id"]
    enq["student_id"] = enq["mobile_clean"].map(smap)

    enq["source_sheet"] = "enquiries"
    enq["possible_duplicate"] = ""

    # `converted` is a set of enquiries the institute already knows converted,
    # and its mobiles live in the ADMISSIONS namespace (124/169 match) rather
    # than the enquiry one (1/169). Folded in as enquiry rows, those 124 are
    # the only enquiry->admission pairs that exist anywhere in the two files.
    conv = conv.copy()
    conv["mobile_clean"] = pd.to_numeric(conv["mobile_clean"], errors="coerce").astype("Int64")
    conv["ENQ_ID"] = [f"CNV-{i:05d}" for i in range(1, len(conv) + 1)]
    conv["course_category"] = [cm.canonicalize_course_category(v) for v in conv["course"]]
    conv["branch"] = conv["branch"].map(
        lambda v: (cm.branch_from_locality(v) or cm.canonicalize_branch(v)))
    conv["branch"] = conv["branch"].map(lambda v: v.title() if isinstance(v, str) else v)
    conv["counsellor"] = conv["counsellor"].map(cm.canonicalize_faculty).map(
        lambda v: v.title() if isinstance(v, str) else v)
    conv["student_id"] = conv["mobile_clean"].map(smap)
    conv["source_sheet"] = "converted"
    # No lead source, education or repeat flag is recorded on this sheet.
    for col in ("lead_source", "lead_source_raw", "education_level",
                "presently_doing", "is_repeat_enquiry"):
        conv[col] = pd.NA
    # 17 names appear in both sheets. Flagged, not merged: a shared name is
    # not proof of a shared person, and the mobiles disagree.
    seen = set(enq["name"].dropna().astype(str).str.lower().str.strip())
    conv["possible_duplicate"] = conv["name"].map(
        lambda v: "name also in enquiries"
        if isinstance(v, str) and v.strip().lower() in seen else "")

    cols = ["ENQ_ID", "enquiry_timestamp", "name", "mobile_clean", "course",
            "course_category", "branch", "counsellor", "mode_of_enquiry",
            "education_level", "presently_doing", "lead_source",
            "lead_source_raw", "is_repeat_enquiry", "student_id",
            "source_sheet", "possible_duplicate"]
    enq = pd.concat([enq[cols], conv[cols]], ignore_index=True)
    bridged = int(enq["student_id"].notna().sum())
    enq = enq.rename(columns={"mobile_clean": "mobile"})
    fu = fu[["ENQ_ID", "call_attempt_number", "call_date",
             "outcome_classified", "discussion_raw"]]
    return enq, fu, repaired, bridged


LOOKUPS = {
    "branch": list(b.title() for b in cm.BRANCHES),
    "course_category": list(cm.COURSE_CATEGORY_BUCKETS),
    "lead_source": list(cm.LEAD_SOURCE_BUCKETS),
    "presently_doing": list(cm.OCCUPATION_BUCKETS),
    "status": PRECEDENCE + ["Unknown"],
    "mode_of_payment": ["Cash", "Online", "Cheque"],
    "mode_of_enquiry": ["Online Form", "On Call"],
}


def write_lookups(writer):
    longest = max(len(v) for v in LOOKUPS.values())
    frame = pd.DataFrame({k: v + [""] * (longest - len(v))
                          for k, v in LOOKUPS.items()})
    frame.to_excel(writer, sheet_name="lookups", index=False)
    return frame


def add_validation(ws, header, lookup_col, n_rows, frame):
    """Dropdown on `header`, pointing at the lookups sheet."""
    if header not in [c.value for c in ws[1]]:
        return
    idx = [c.value for c in ws[1]].index(header) + 1
    col = get_column_letter(idx)
    lk = get_column_letter(list(frame.columns).index(lookup_col) + 1)
    n = len([v for v in LOOKUPS[lookup_col] if v])
    dv = DataValidation(type="list",
                        formula1=f"=lookups!${lk}$2:${lk}${n + 1}",
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{n_rows + 1}")


def add_lookup_column(ws, header, formula_template, n_rows):
    """Append a live XLOOKUP column that reads from the parent sheet."""
    idx = ws.max_column + 1
    ws.cell(row=1, column=idx, value=header)
    for r in range(2, n_rows + 2):
        ws.cell(row=r, column=idx, value=formula_template.format(row=r))


def main(src: str = "", out: str = "") -> None:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    src, out = src or SRC, out or OUT
    os.makedirs(out, exist_ok=True)
    adm, st, fee, cert, enq, fu, conv = (
        read(n, src) for n in ("admissions", "status", "fee_receipts",
                               "certificates", "enquiries", "follow_up_calls",
                               "converted"))

    students, from_status, from_cert = build_students(adm, st, fee, cert)
    enq2, fu2, repaired, bridged = repair_enquiries(enq, fu, conv, students)

    # Child tables keep only what belongs at their own grain. course/branch/
    # category live on the master; duplicated here they would drift out of
    # agreement with it.
    fee2 = fee.copy()
    fee2["student_id"] = as_id(fee2["student_id"])
    fee2 = fee2[["receipt_id", "student_id", "date_of_receipt", "paid_amt",
                 "mode_of_payment"]]
    cert2 = cert.copy()
    cert2["student_id"] = as_id(cert2["student_id"])
    cert2 = cert2[["student_id", "certificate_number", "certificate_issue_date"]]

    # ---------------------------------------------------- students workbook
    path_b = os.path.join(out, "FV_Students_v2.xlsx")
    with pd.ExcelWriter(path_b, engine="openpyxl") as w:
        students.to_excel(w, sheet_name="students", index=False)
        fee2.to_excel(w, sheet_name="fee_receipts", index=False)
        cert2.to_excel(w, sheet_name="certificates", index=False)
        frame = write_lookups(w)

        ws = w.book["students"]
        for header, lookup in (("branch", "branch"),
                               ("course_category", "course_category"),
                               ("lead_source", "lead_source"),
                               ("presently_doing", "presently_doing"),
                               ("status", "status")):
            add_validation(ws, header, lookup, len(students), frame)

        n = len(students)
        rng = f"students!$A$2:$A${n + 1}"
        for sheet, rows in (("fee_receipts", len(fee2)),
                            ("certificates", len(cert2))):
            child = w.book[sheet]
            add_lookup_column(
                child, "student_name",
                f'=IFERROR(XLOOKUP(B{{row}},{rng},students!$B$2:$B${n + 1}),"")',
                rows)
            add_lookup_column(
                child, "branch_ref",
                f'=IFERROR(XLOOKUP(B{{row}},{rng},students!$E$2:$E${n + 1}),"")',
                rows)
            add_validation(child, "mode_of_payment", "mode_of_payment",
                           rows, frame)

    # ---------------------------------------------------- enquiry workbook
    path_a = os.path.join(out, "FV_Enquiry_v2.xlsx")
    with pd.ExcelWriter(path_a, engine="openpyxl") as w:
        enq2.to_excel(w, sheet_name="enquiries", index=False)
        fu2.to_excel(w, sheet_name="follow_up_calls", index=False)
        frame = write_lookups(w)

        ws = w.book["enquiries"]
        for header, lookup in (("branch", "branch"),
                               ("course_category", "course_category"),
                               ("lead_source", "lead_source"),
                               ("presently_doing", "presently_doing"),
                               ("mode_of_enquiry", "mode_of_enquiry")):
            add_validation(ws, header, lookup, len(enq2), frame)

        m = len(enq2)
        rng = f"enquiries!$A$2:$A${m + 1}"
        child = w.book["follow_up_calls"]
        add_lookup_column(
            child, "enquirer_name",
            f'=IFERROR(XLOOKUP(A{{row}},{rng},enquiries!$C$2:$C${m + 1}),"")',
            len(fu2))
        add_lookup_column(
            child, "branch_ref",
            f'=IFERROR(XLOOKUP(A{{row}},{rng},enquiries!$G$2:$G${m + 1}),"")',
            len(fu2))

    # ------------------------------------------------------------- report
    total = len(students)
    unknown = int((students["status"] == "Unknown").sum())
    print(f"students master           {total:5d} rows")
    print(f"  status from status sheet{from_status:5d}")
    print(f"  status from certificate {from_cert:5d}")
    print(f"  still Unknown           {unknown:5d}  ({100 * unknown / total:.0f}%)")
    print(f"\nstatus breakdown:")
    for k, v in students["status"].value_counts().items():
        print(f"    {k:20s}{v:5d}")
    print(f"\nenquiries                {len(enq2):5d} rows")
    print(f"  follow-up calls given a real ENQ_ID {repaired}/{len(fu2)}")
    print(f"  bridged to a student_id             {bridged}/{len(enq2)}")
    print(f"\ncourse_category coverage {students['course_category'].notna().sum()}/{total}")
    print(f"branch coverage          {students['branch'].notna().sum()}/{total}")

    # How much of the Unknown backlog is recent enough to be worth chasing.
    unk = students[students["status"] == "Unknown"].copy()
    joined = pd.to_datetime(unk["joining_date"], errors="coerce", dayfirst=True)
    print(f"\nUNKNOWN STATUS BACKLOG ({len(unk)}), by joining year:")
    by_year = joined.dt.year.value_counts(dropna=False).sort_index()
    for yr, n in by_year.items():
        label = "no joining date" if pd.isna(yr) else str(int(yr))
        print(f"    {label:18s}{n:5d}")
    recent = int((joined >= "2024-01-01").sum())
    print(f"  joined 2024 or later     {recent:5d}  <- chase these first")
    print(f"  older / undated          {len(unk) - recent:5d}  probably long finished")
    print(f"\nwrote {path_b}")
    print(f"wrote {path_a}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--src", default=SRC,
                    help="directory of exported sheet CSVs")
    ap.add_argument("--out", default=OUT,
                    help="directory the two workbooks are written to")
    args = ap.parse_args()
    main(args.src, args.out)
